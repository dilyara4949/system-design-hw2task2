import os
import psycopg2
import zipfile
from config import db_config
import openpyxl  # You may need to install this library if not already installed

# Step 1: Unzip the Excel file
zip_file_path = "selection_data.zip"
extracted_folder = "extracted_data"

with zipfile.ZipFile(zip_file_path, 'r') as zip_ref:
    zip_ref.extractall(extracted_folder)

# Assuming there's a single Excel file in the extracted folder
excel_file_name = os.listdir(extracted_folder)[0]
excel_file_path = os.path.join(extracted_folder, excel_file_name)

# Step 2: Read data from the Excel file
workbook = openpyxl.load_workbook(excel_file_path)
worksheet = workbook.active  # Assuming the data is in the active sheet


conn = psycopg2.connect(**db_config)

# Step 4: Create a cursor object
cur = conn.cursor()


# Get the column names from the first row of the Excel file
# header_row = next(worksheet.iter_rows(min_row=1, max_row=1, values_only=True))
header_row = worksheet[2]
column_names = [str(col_name.value) for col_name in header_row]



# Create the SQL query to create the table
table_name = "TASK"  # Replace with your desired table name
create_table_sql = f"CREATE TABLE IF NOT EXISTS {table_name} ("
for column_name in column_names:
    create_table_sql += f'"{column_name}" TEXT,'  # Assume all columns are of type TEXT, you can adjust as needed
create_table_sql = create_table_sql.rstrip(",") + ");"  # Remove the trailing comma and add the closing parenthesis
cur.execute(create_table_sql)
conn.commit()



# Insert data into the table
for row in worksheet.iter_rows(min_row=2, values_only=True):
    placeholders = ",".join(["%s"] * len(row))
    insert_sql = f"INSERT INTO {table_name} VALUES ({placeholders});"
    cur.execute(insert_sql, row)

# Step 6: Commit the changes and close the cursor and connection
conn.commit()
cur.close()
conn.close()
